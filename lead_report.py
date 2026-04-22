from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
import pandas as pd
import re
import time
from datetime import datetime
import os

# ════════════════════════════
# STEP 1 — SCRAPE
# ════════════════════════════
def scrape(search_query, max_results=50):
    results = []

    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-notifications")
    options.add_argument("--lang=en")

    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options
    )

    try:
        query = search_query.replace(" ", "+")
        driver.get(
            f"https://www.google.com/maps/search/{query}/?hl=en"
        )
        time.sleep(6)

        # Handle popup
        try:
            agree = WebDriverWait(driver, 6).until(
                EC.element_to_be_clickable((By.XPATH,
                '//button[contains(., "Accept") or contains(., "Agree")]'
                ))
            )
            agree.click()
            time.sleep(2)
        except:
            pass

        # Scroll to load all results
        print("📜 Loading results...")
        for i in range(15):
            try:
                scrollable = driver.find_element(
                    By.XPATH, '//div[@role="feed"]'
                )
                driver.execute_script(
                    "arguments[0].scrollTop += 1500", scrollable
                )
                time.sleep(3)
                end = driver.find_elements(By.XPATH,
                    '//*[contains(text(),"end of the list")]'
                )
                if end:
                    break
            except:
                time.sleep(2)
            print(f"Scroll {i+1}/15")

        # Collect all links
        links = driver.find_elements(
            By.XPATH, '//a[contains(@href, "/maps/place/")]'
        )

        listing_data = []
        seen = set()
        for link in links:
            name = link.get_attribute("aria-label")
            href = link.get_attribute("href")
            if name and href and name not in seen:
                seen.add(name)
                listing_data.append({"name": name, "href": href})

        print(f"\n✅ Found {len(listing_data)} businesses\n")

        # Visit each business
        for item in listing_data:
            if len(results) >= max_results:
                break
            try:
                driver.get(item["href"])
                time.sleep(3)

                try:
                    phone = driver.find_element(By.XPATH,
                        '//button[contains(@data-item-id,"phone")]//div[contains(@class,"fontBodyMedium")]'
                    ).text
                except:
                    phone = "N/A"

                try:
                    address = driver.find_element(By.XPATH,
                        '//button[@data-item-id="address"]//div[contains(@class,"fontBodyMedium")]'
                    ).text
                except:
                    address = "N/A"

                try:
                    rating = driver.find_element(By.XPATH,
                        '//span[@aria-hidden="true" and contains(@class,"ceNzKf") or @role="img"]'
                    ).get_attribute("aria-label")
                except:
                    rating = "N/A"

                try:
                    website = driver.find_element(By.XPATH,
                        '//a[@data-item-id="authority"]'
                    ).get_attribute("href")
                except:
                    website = "N/A"

                results.append({
                    "Business Name": item["name"],
                    "Phone": phone,
                    "Address": address,
                    "Rating": rating,
                    "Website": website
                })
                print(f"✅ {len(results)}. {item['name']} | {phone}")

            except Exception as e:
                continue

    finally:
        driver.quit()

    return results


# ════════════════════════════
# STEP 2 — CLEAN
# ════════════════════════════
def clean(data):
    df = pd.DataFrame(data)

    # Clean phone
    def clean_phone(phone):
        if phone == "N/A" or pd.isna(phone):
            return "N/A"
        digits = re.sub(r'\D', '', str(phone))
        if len(digits) == 10:
            return f"+91 {digits[:5]} {digits[5:]}"
        elif len(digits) == 12 and digits.startswith("91"):
            digits = digits[2:]
            return f"+91 {digits[:5]} {digits[5:]}"
        return str(phone)

    # Clean rating
    def clean_rating(rating):
        if rating == "N/A" or pd.isna(rating):
            return "N/A"
        try:
            match = re.search(r'(\d+\.?\d*)', str(rating))
            if match:
                return float(match.group(1))
        except:
            pass
        return "N/A"

    # Lead quality
    def lead_quality(row):
        score = 0
        if row.get("Phone", "N/A") != "N/A":
            score += 1
        if row.get("Website", "N/A") != "N/A":
            score += 1
        if row.get("Rating", "N/A") != "N/A":
            try:
                if float(row["Rating"]) >= 4.0:
                    score += 1
            except:
                pass
        if score == 3:
            return "Hot"
        elif score == 2:
            return "Good"
        else:
            return "Cold"
        
    df.drop_duplicates(subset=["Business Name"], inplace=True)
    df["Phone"] = df["Phone"].apply(clean_phone)
    df["Rating"] = df["Rating"].apply(clean_rating)
    df = df.fillna("N/A")
    df = df.replace("", "N/A")
    df = df.replace(" ", "N/A")
    df = df.replace(r'^\s*$', "N/A", regex=True)
    df["Lead Quality"] = df.apply(lead_quality, axis=1)    
        


    # Sort by rating
    df_rated = df[df["Rating"] != "N/A"].copy()
    df_no_rating = df[df["Rating"] == "N/A"].copy()
    try:
        df_rated = df_rated.sort_values(by="Rating", ascending=False)
    except:
        pass
    df = pd.concat([df_rated, df_no_rating], ignore_index=True)

    print(f"\n🧹 Cleaned {len(df)} leads")
    return df


# ════════════════════════════
# STEP 3 — GENERATE REPORT
# ════════════════════════════
def generate_report(df, search_query):
    date = datetime.now().strftime("%Y-%m-%d")
    filename = f"lead_report_{date}.xlsx"

    wb = openpyxl.Workbook()

    # ── Sheet 1: All Leads ──
    ws1 = wb.active
    ws1.title = "All Leads"

    headers = list(df.columns)
    ws1.append(headers)

    for cell in ws1[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor="1565C0")
        cell.alignment = Alignment(horizontal="center")

    for idx, row in df.iterrows():
        ws1.append(list(row))
        row_num = idx + 2
        color = "E3F2FD" if idx % 2 == 0 else "FFFFFF"
        for cell in ws1[row_num]:
            cell.fill = PatternFill("solid", fgColor=color)

    for col in ws1.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws1.column_dimensions[
            col[0].column_letter
        ].width = min(max_len + 4, 50)

    # ── Sheet 2: Top 10 ──
    ws2 = wb.create_sheet("Top 10 Leads")
    top10 = df[df["Phone"] != "N/A"].head(10)

    ws2.append(headers)
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor="B71C1C")
        cell.alignment = Alignment(horizontal="center")

    for idx, row in top10.iterrows():
        ws2.append(list(row))

    for col in ws2.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws2.column_dimensions[
            col[0].column_letter
        ].width = min(max_len + 4, 50)

    # ── Sheet 3: Summary ──
    ws3 = wb.create_sheet("Summary Report")

    total = len(df)
    with_phone = len(df[df["Phone"] != "N/A"])
    with_website = len(df[df["Website"] != "N/A"])
    hot = len(df[df["Lead Quality"] == "Hot"])
    good = len(df[df["Lead Quality"] == "Good"])
    cold = len(df[df["Lead Quality"] == "Cold"])

    try:
        avg_rating = round(
            float(df[df["Rating"] != "N/A"]["Rating"].mean()), 2
        )
    except:
        avg_rating = "N/A"

    # Title
    ws3["A1"] = "AUTOMATED LEAD GENERATION REPORT"
    ws3["A1"].font = Font(bold=True, size=16, color="1565C0")
    ws3["A2"] = f"Search: {search_query}"
    ws3["A3"] = f"Generated: {datetime.now().strftime('%d %B %Y %I:%M %p')}"
    ws3["A3"].font = Font(italic=True, color="666666")

    ws3.append([])

    # Stats
    summary = [
        ["METRIC", "VALUE"],
        ["Total Leads Scraped", total],
        ["Leads With Phone", with_phone],
        ["Leads With Website", with_website],
        ["Average Rating", avg_rating],
        [],
        ["LEAD QUALITY", "COUNT"],
        ["Hot Leads", hot],
        ["Good Leads", good],
        ["Cold Leads", cold],
    ]

    for row in summary:
        ws3.append(row)

    # Style headers
    for row_num in [5, 11]:
        ws3.cell(row=row_num, column=1).font = Font(
            bold=True, color="FFFFFF"
        )
        ws3.cell(row=row_num, column=1).fill = PatternFill(
            "solid", fgColor="1565C0"
        )
        ws3.cell(row=row_num, column=2).font = Font(
            bold=True, color="FFFFFF"
        )
        ws3.cell(row=row_num, column=2).fill = PatternFill(
            "solid", fgColor="1565C0"
        )

    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 20

    # Add bar chart
    chart = BarChart()
    chart.title = "Lead Quality Breakdown"
    chart.style = 10
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Quality"

    data = Reference(ws3, min_col=2, min_row=12, max_row=14)
    cats = Reference(ws3, min_col=1, min_row=12, max_row=14)
    chart.add_data(data)
    chart.set_categories(cats)
    chart.shape = 4
    ws3.add_chart(chart, "D5")

    wb.save(filename)
    print(f"\n💾 Report saved: {filename}")
    return filename


# ════════════════════════════
# MAIN — RUN EVERYTHING
# ════════════════════════════
if __name__ == "__main__":
    print("🚀 AUTOMATED LEAD GENERATION REPORT")
    print("=" * 45)

    SEARCH = "digital marketing agencies in Chennai"
    MAX = 50

    print(f"\n🔍 Search: {SEARCH}")
    print(f"🎯 Target: {MAX} leads\n")

    # Step 1
    print("STEP 1 — SCRAPING...")
    print("-" * 30)
    data = scrape(SEARCH, MAX)

    # Step 2
    print("\nSTEP 2 — CLEANING...")
    print("-" * 30)
    df = clean(data)

    # Step 3
    print("\nSTEP 3 — GENERATING REPORT...")
    print("-" * 30)
    report = generate_report(df, SEARCH)

    print(f"\n{'='*45}")
    print(f"🎉 DONE! Full report ready: {report}")
    print(f"{'='*45}")